from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, extract
from typing import Optional
from datetime import date, datetime
import io

from app.db.database import get_db
from app.models.models import (
    Supplier, Category, Expense, ExpenseCategory,
    Product, Sale, SaleItem, User
)
from app.schemas.schemas import (
    SupplierCreate, SupplierUpdate, SupplierOut,
    CategoryCreate, CategoryOut,
    ExpenseCreate, ExpenseUpdate, ExpenseOut,
    ExpenseCategoryCreate, ExpenseCategoryOut,
    DashboardStats, SaleOut
)
from app.core.deps import get_current_user, admin_or_manager

# ─── Suppliers ────────────────────────────────────────────────────────────────
supplier_router = APIRouter(prefix="/suppliers", tags=["Suppliers"])


@supplier_router.get("", response_model=dict)
def list_suppliers(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Supplier)
    if search:
        q = q.filter(Supplier.name.ilike(f"%{search}%") | Supplier.company_name.ilike(f"%{search}%"))
    total = q.count()
    items = q.offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": [SupplierOut.model_validate(s) for s in items],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@supplier_router.post("", response_model=SupplierOut, status_code=201)
def create_supplier(payload: SupplierCreate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    s = Supplier(**payload.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


@supplier_router.get("/{supplier_id}", response_model=SupplierOut)
def get_supplier(supplier_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return s


@supplier_router.put("/{supplier_id}", response_model=SupplierOut)
def update_supplier(supplier_id: int, payload: SupplierUpdate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(s, field, value)
    db.commit()
    db.refresh(s)
    return s


@supplier_router.delete("/{supplier_id}")
def delete_supplier(supplier_id: int, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    s.is_active = False
    db.commit()
    return {"message": "Supplier deactivated"}


# ─── Categories ───────────────────────────────────────────────────────────────
category_router = APIRouter(prefix="/categories", tags=["Categories"])


@category_router.get("", response_model=list[CategoryOut])
def list_categories(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(Category).all()


@category_router.post("", response_model=CategoryOut, status_code=201)
def create_category(payload: CategoryCreate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    if db.query(Category).filter(Category.name == payload.name).first():
        raise HTTPException(status_code=400, detail="Category already exists")
    c = Category(**payload.model_dump())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@category_router.put("/{cat_id}", response_model=CategoryOut)
def update_category(cat_id: int, payload: CategoryCreate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    c = db.query(Category).filter(Category.id == cat_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    c.name = payload.name
    c.description = payload.description
    db.commit()
    db.refresh(c)
    return c


@category_router.delete("/{cat_id}")
def delete_category(cat_id: int, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    c = db.query(Category).filter(Category.id == cat_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    db.delete(c)
    db.commit()
    return {"message": "Category deleted"}


# ─── Expense Categories ───────────────────────────────────────────────────────
expense_cat_router = APIRouter(prefix="/expense-categories", tags=["Expense Categories"])


@expense_cat_router.get("", response_model=list[ExpenseCategoryOut])
def list_expense_cats(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(ExpenseCategory).all()


@expense_cat_router.post("", response_model=ExpenseCategoryOut, status_code=201)
def create_expense_cat(payload: ExpenseCategoryCreate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    ec = ExpenseCategory(**payload.model_dump())
    db.add(ec)
    db.commit()
    db.refresh(ec)
    return ec


# ─── Expenses ─────────────────────────────────────────────────────────────────
expense_router = APIRouter(prefix="/expenses", tags=["Expenses"])


@expense_router.get("", response_model=dict)
def list_expenses(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    category_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(admin_or_manager),
):
    q = db.query(Expense).options(joinedload(Expense.category))
    if category_id:
        q = q.filter(Expense.category_id == category_id)
    if date_from:
        q = q.filter(func.date(Expense.expense_date) >= date_from)
    if date_to:
        q = q.filter(func.date(Expense.expense_date) <= date_to)
    total = q.count()
    items = q.order_by(Expense.expense_date.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": [ExpenseOut.model_validate(e) for e in items],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@expense_router.post("", response_model=ExpenseOut, status_code=201)
def create_expense(payload: ExpenseCreate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    e = Expense(**payload.model_dump())
    db.add(e)
    db.commit()
    db.refresh(e)
    return db.query(Expense).options(joinedload(Expense.category)).filter(Expense.id == e.id).first()


@expense_router.put("/{expense_id}", response_model=ExpenseOut)
def update_expense(expense_id: int, payload: ExpenseUpdate, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    e = db.query(Expense).filter(Expense.id == expense_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Expense not found")
    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(e, field, value)
    db.commit()
    db.refresh(e)
    return e


@expense_router.delete("/{expense_id}")
def delete_expense(expense_id: int, db: Session = Depends(get_db), _=Depends(admin_or_manager)):
    e = db.query(Expense).filter(Expense.id == expense_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Expense not found")
    db.delete(e)
    db.commit()
    return {"message": "Expense deleted"}


# ─── Dashboard ────────────────────────────────────────────────────────────────
dashboard_router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


@dashboard_router.get("/stats")
def get_dashboard_stats(db: Session = Depends(get_db), _=Depends(get_current_user)):
    today = date.today()
    now = datetime.now()

    total_products = db.query(Product).filter(Product.is_active == True).count()
    total_suppliers = db.query(Supplier).filter(Supplier.is_active == True).count()

    today_sales = db.query(Sale).filter(func.date(Sale.created_at) == today).all()
    revenue_today = sum(s.total_amount for s in today_sales)

    month_sales = db.query(Sale).filter(
        extract("year", Sale.created_at) == now.year,
        extract("month", Sale.created_at) == now.month
    ).all()
    revenue_month = sum(s.total_amount for s in month_sales)

    cost_month = sum(
        item.cost_price * item.quantity
        for s in month_sales
        for item in s.items
    )

    expenses_month = db.query(func.sum(Expense.amount)).filter(
        extract("year", Expense.expense_date) == now.year,
        extract("month", Expense.expense_date) == now.month
    ).scalar() or 0.0

    low_stock = db.query(Product).filter(
        Product.quantity <= Product.low_stock_threshold,
        Product.quantity > 0,
        Product.is_active == True
    ).count()
    out_of_stock = db.query(Product).filter(
        Product.quantity == 0,
        Product.is_active == True
    ).count()

    recent_sales = db.query(Sale).options(
        joinedload(Sale.cashier),
        joinedload(Sale.items).joinedload(SaleItem.product)
    ).order_by(Sale.created_at.desc()).limit(5).all()

    low_stock_products = db.query(Product).options(
        joinedload(Product.category)
    ).filter(
        Product.quantity <= Product.low_stock_threshold,
        Product.is_active == True
    ).order_by(Product.quantity).limit(10).all()

    monthly_revenue = []
    for month in range(1, 13):
        rev = db.query(func.sum(Sale.total_amount)).filter(
            extract("year", Sale.created_at) == now.year,
            extract("month", Sale.created_at) == month
        ).scalar() or 0.0
        exp = db.query(func.sum(Expense.amount)).filter(
            extract("year", Expense.expense_date) == now.year,
            extract("month", Expense.expense_date) == month
        ).scalar() or 0.0
        monthly_revenue.append({
            "month": month,
            "revenue": round(rev, 2),
            "expenses": round(exp, 2),
        })

    top_products_raw = db.query(
        SaleItem.product_id,
        func.sum(SaleItem.quantity).label("sold"),
        func.sum(SaleItem.total).label("revenue")
    ).group_by(SaleItem.product_id).order_by(
        func.sum(SaleItem.quantity).desc()
    ).limit(5).all()

    top_products = []
    for row in top_products_raw:
        p = db.query(Product).filter(Product.id == row.product_id).first()
        if p:
            top_products.append({
                "name": p.name,
                "sold": row.sold,
                "revenue": round(row.revenue, 2),
            })

    return {
        "total_products": total_products,
        "total_sales_today": len(today_sales),
        "revenue_today": round(revenue_today, 2),
        "revenue_month": round(revenue_month, 2),
        "low_stock_count": low_stock,
        "out_of_stock_count": out_of_stock,
        "total_expenses_month": round(expenses_month, 2),
        "profit_month": round(revenue_month - cost_month - expenses_month, 2),
        "total_suppliers": total_suppliers,
        "recent_sales": [SaleOut.model_validate(s) for s in recent_sales],
        "low_stock_products": low_stock_products,
        "monthly_revenue": monthly_revenue,
        "top_products": top_products,
    }


# ─── Reports ──────────────────────────────────────────────────────────────────
reports_router = APIRouter(prefix="/reports", tags=["Reports"])


@reports_router.get("/sales")
def sales_report(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(admin_or_manager),
):
    q = db.query(Sale)
    if date_from:
        q = q.filter(func.date(Sale.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(Sale.created_at) <= date_to)
    sales = q.all()
    return {
        "total_sales": len(sales),
        "total_revenue": round(sum(s.total_amount for s in sales), 2),
        "total_discount": round(sum(s.discount_amount for s in sales), 2),
        "by_payment_method": {
            method: {"count": 0, "total": 0.0}
            for method in set(s.payment_method for s in sales)
        },
    }


@reports_router.get("/export/excel")
def export_excel(
    report_type: str = "sales",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(admin_or_manager),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.capitalize()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")

    if report_type == "sales":
        headers = ["Reference", "Date", "Cashier", "Items", "Subtotal", "Discount", "Total", "Payment", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        q = db.query(Sale).options(joinedload(Sale.cashier), joinedload(Sale.items))
        if date_from:
            q = q.filter(func.date(Sale.created_at) >= date_from)
        if date_to:
            q = q.filter(func.date(Sale.created_at) <= date_to)

        for row_num, sale in enumerate(q.all(), 2):
            ws.cell(row=row_num, column=1, value=sale.reference)
            ws.cell(row=row_num, column=2, value=sale.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row_num, column=3, value=sale.cashier.full_name if sale.cashier else "")
            ws.cell(row=row_num, column=4, value=len(sale.items))
            ws.cell(row=row_num, column=5, value=sale.subtotal)
            ws.cell(row=row_num, column=6, value=sale.discount_amount)
            ws.cell(row=row_num, column=7, value=sale.total_amount)
            ws.cell(row=row_num, column=8, value=sale.payment_method)
            ws.cell(row=row_num, column=9, value=sale.status.value)

    elif report_type == "products":
        headers = ["ID", "Name", "SKU", "Category", "Supplier", "Cost Price", "Selling Price", "Quantity", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        products = db.query(Product).options(
            joinedload(Product.category),
            joinedload(Product.supplier)
        ).all()
        for row_num, p in enumerate(products, 2):
            ws.cell(row=row_num, column=1, value=p.id)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.sku)
            ws.cell(row=row_num, column=4, value=p.category.name if p.category else "")
            ws.cell(row=row_num, column=5, value=p.supplier.name if p.supplier else "")
            ws.cell(row=row_num, column=6, value=p.cost_price)
            ws.cell(row=row_num, column=7, value=p.selling_price)
            ws.cell(row=row_num, column=8, value=p.quantity)
            ws.cell(row=row_num, column=9, value="Active" if p.is_active else "Inactive")

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"},
    )


@reports_router.get("/export/pdf")
def export_pdf(
    report_type: str = "sales",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(admin_or_manager),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"StockPilot - {report_type.capitalize()} Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    if report_type == "sales":
        q = db.query(Sale).options(joinedload(Sale.cashier))
        if date_from:
            q = q.filter(func.date(Sale.created_at) >= date_from)
        if date_to:
            q = q.filter(func.date(Sale.created_at) <= date_to)
        sales = q.all()

        data = [["Reference", "Date", "Cashier", "Total", "Payment", "Status"]]
        for s in sales:
            data.append([
                s.reference,
                s.created_at.strftime("%Y-%m-%d"),
                s.cashier.full_name if s.cashier else "",
                f"${s.total_amount:.2f}",
                s.payment_method,
                s.status.value,
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={report_type}_report.pdf"},
    )